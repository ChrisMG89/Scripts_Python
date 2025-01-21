import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import customtkinter as ctk
from tkinter import messagebox, filedialog


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Generador de Excel de Vuelos")
        self.geometry("600x300")
        self.iconbitmap(default='combined_icon.ico')

        # Elementos de la interfaz
        self.label = ctk.CTkLabel(self, text="Selecciona el directorio de las miniaturas:")
        self.label.pack(pady=10)

        self.directory_entry = ctk.CTkEntry(self, width=400)
        self.directory_entry.pack(pady=10)

        self.browse_button = ctk.CTkButton(self, text="Buscar", command=self.browse_directory)
        self.browse_button.pack(pady=10)

        self.generate_button = ctk.CTkButton(self, text="Generar Excel", command=self.generate_excel)
        self.generate_button.pack(pady=10)

    def browse_directory(self):
        """Abre un cuadro de diálogo para seleccionar un directorio."""
        directory = filedialog.askdirectory()
        if directory:
            self.directory_entry.delete(0, 'end')
            self.directory_entry.insert(0, directory)

    def obtener_nombres_vuelos(self, directorio):
        """Obtiene los nombres de las carpetas dentro de un directorio."""
        carpetas = [nombre for nombre in os.listdir(directorio) if os.path.isdir(os.path.join(directorio, nombre))]
        nombres_vuelos = [nombre.replace('_miniaturas', '') for nombre in carpetas]
        return nombres_vuelos

    def obtener_datos_directorio(self, directorio):
        """Extrae el nombre de la planta y el año del directorio."""
        partes_directorio = os.path.normpath(directorio).split(os.sep)
        nombre_planta = partes_directorio[8].upper()  # Convertir a mayúsculas
        año_inspeccion = partes_directorio[9]
        return nombre_planta, año_inspeccion

    def generar_excel_vuelos(self, directorio, nombre_archivo_excel, nombre_planta):
        """Genera un archivo Excel con los datos de los vuelos."""
        vuelos = self.obtener_nombres_vuelos(directorio)

        columnas_obligatorias = ["Vuelo", "Extraccion", "Renom", "Comprim", "GEO+CSV", "Girado",
                                 "Miniaturas", "PROCESADO", "AWS", "INSPECC", "EXTRACC", "TIFF",
                                 "THERMAL", "INFORMES"]

        # Crear un DataFrame vacío con las columnas obligatorias
        df = pd.DataFrame(columns=columnas_obligatorias)
        df["Vuelo"] = vuelos

        # Guardar inicialmente el DataFrame en un archivo Excel
        df.to_excel(nombre_archivo_excel, index=False, startrow=1)

        # Cargar el archivo con openpyxl para poder formatearlo
        wb = load_workbook(nombre_archivo_excel)
        ws = wb.active

        # Fusionar celdas B1 y C1, y colocar el nombre de la planta
        ws.merge_cells('B1:C1')
        ws['B1'] = nombre_planta
        ws['B1'].font = Font(name='Calibri', bold=True, size=16)
        ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B1'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Estilo para los encabezados
        fill_column = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        font_column = Font(name='Calibri', bold=True, size=11)

        # Ajustar el ancho de las columnas (más pequeño)
        for col_num, column_title in enumerate(columnas_obligatorias, 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 15  # Ajustar ancho (más pequeño)

            # Establecer el formato de los encabezados
            cell = ws.cell(row=2, column=col_num)
            cell.value = column_title
            cell.fill = fill_column
            cell.font = font_column
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Definir los bordes de las celdas
        medium_border = Border(left=Side(style='medium'), right=Side(style='medium'),
                               top=Side(style='medium'), bottom=Side(style='medium'))
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # Aplicar bordes a los encabezados
        for col in range(1, len(columnas_obligatorias) + 1):
            cell = ws.cell(row=2, column=col)
            # Eliminar bordes internos entre B2 y N2
            if col == 1:
                cell.border = Border(left=Side(style='medium'), right=Side(style='none'),
                                     top=Side(style='medium'), bottom=Side(style='medium'))
            elif col == len(columnas_obligatorias):
                cell.border = Border(left=Side(style='none'), right=Side(style='medium'),
                                     top=Side(style='medium'), bottom=Side(style='medium'))
            else:
                cell.border = Border(left=Side(style='none'), right=Side(style='none'),
                                     top=Side(style='medium'), bottom=Side(style='medium'))

        # Aplicar bordes y alineación a las filas de vuelos (sin eliminar bordes interiores)
        for row in range(3, len(vuelos) + 3):
            for col in range(1, len(columnas_obligatorias) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Guardar el archivo Excel con el formato aplicado
        wb.save(nombre_archivo_excel)
        messagebox.showinfo("Éxito", f"Archivo Excel '{nombre_archivo_excel}' generado correctamente.")

    def generate_excel(self):
        """Genera el archivo Excel basado en el directorio ingresado."""
        directorio_vuelos = self.directory_entry.get()
        if not os.path.isdir(directorio_vuelos):
            messagebox.showerror("Error", "El directorio especificado no es válido.")
            return

        nombre_planta, año_inspeccion = self.obtener_datos_directorio(directorio_vuelos)
        directorio_principal = os.path.dirname(directorio_vuelos)
        nombre_archivo_excel = os.path.join(directorio_principal, f"CONTROL_{nombre_planta}_{año_inspeccion}.xlsx")

        self.generar_excel_vuelos(directorio_vuelos, nombre_archivo_excel, nombre_planta)


if __name__ == "__main__":
    ctk.set_appearance_mode("dark")  # Puedes cambiar a "light" si lo prefieres
    ctk.set_default_color_theme("blue")  # Elige entre "blue" o "dark-blue"

    app = App()
    app.mainloop()
