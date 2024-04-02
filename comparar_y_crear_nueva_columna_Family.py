'''------------------------------------------------------------------------------
Nombre : Comparar y crear nueva columna
Readme:
    Este script de Python combina datos de dos archivos Excel y agrega una columna adicional al primer archivo Excel. 
    El script utiliza las bibliotecas pandas y openpyxl para cargar los datos de los archivos Excel y manipular los libros de trabajo.
Requisitos:
    -Python 3.x
    -Bibliotecas Python: pandas, openpyxl
    -Instalación de dependencias (Puedes instalar las bibliotecas requeridas utilizando pip:)

                                    pip install pandas openpyxl
Uso:
    Abre el script Python en tu entorno de desarrollo o editor de texto preferido.
    Especifica las rutas de los archivos Excel que deseas combinar. Modifica las variables excel1_path y excel2_path con las rutas correctas de tus archivos Excel.
    Ejecuta el script Python. Los datos del primer y segundo archivo Excel se cargarán y fusionarán en un nuevo DataFrame.
    Se agrega una nueva columna llamada "Family" al primer archivo Excel. La columna se agrega en la hoja "Hoja1" del primer archivo Excel.
    Guarda los cambios en el primer archivo Excel.

Notas:
    Asegúrate de tener los archivos Excel existentes antes de ejecutar el script.
    Los archivos Excel deben tener una hoja de datos llamada "Hoja1" para que el script funcione correctamente.
    Si deseas agregar la columna "Family" a una hoja diferente del primer archivo Excel, puedes modificar el script para especificar el nombre de la hoja correcta

----------------------------------------------------------------------------'''

import pandas as pd
from openpyxl import load_workbook

# Especificar las rutas de los archivos Excel
excel1_path = r'C:\Users\Aerotools\Desktop\Nueva carpeta\NUMEROS_SERIE.xlsx'
excel2_path = r'D:\AEROTOOLS-UAV Dropbox\AEROTOOLS-UAV\Proyectos\PLANTAS_PV\SONNEDIX\CASTELNAU\2024\NUMEROS_SERIE\SN_Report.xlsx'

print("Cargando datos del primer Excel...")
# Cargar los datos del primer Excel
df_snor = pd.read_excel(excel1_path, sheet_name='Hoja1')
print("Datos del primer Excel cargados:")
print(df_snor.head())

print("Cargando datos del segundo Excel...")
# Cargar los datos del segundo Excel
df_fr = pd.read_excel(excel2_path)
print("\nDatos del segundo Excel cargados:")
print(df_fr.head())

print("Combinando datos...")
# Fusionar los dos dataframes en base al número de serie
merged_df = pd.merge(df_snor, df_fr, how='left', left_on='Serial Number', right_on='SN Report')
print("\nDatos fusionados:")
print(merged_df.head())

# Agregar la columna 'Family' al primer DataFrame (df_snor)
df_snor['Family'] = merged_df['Family']

# Abrir el libro de trabajo existente
wb = load_workbook(excel1_path)

# Obtener la hoja existente
sheet = wb['Hoja1']

# Obtener la letra de la columna siguiente a la última columna ocupada
next_column_letter = chr(sheet.max_column + 1 + 64)

# Agregar el encabezado "Family" en la nueva columna
sheet[f'{next_column_letter}1'] = "Family"

# Asignar valores de la columna 'Family' a la columna siguiente a la última columna ocupada
for i, value in enumerate(df_snor['Family'], start=2):  # Comenzamos desde la fila 2 para dejar espacio para el encabezado
    sheet[f'{next_column_letter}{i}'] = value

# Guardar el libro de trabajo
wb.save(excel1_path)

print("La columna 'Family' ha sido agregada al primer Excel en la hoja 'Hoja1'.")
